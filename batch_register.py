#!/usr/bin/env python3
"""
批量机场面板自动注册 + 订阅提取脚本 (v2 - 全自动优化版)
修复: CF绕过、SPA面板探测、验证码IMAP、SSPanel字段、超时快速失败
"""

import cloudscraper
import requests
import json
import time
import re
import imaplib
import email
from email.header import decode_header
from urllib.parse import urlparse, parse_qs, urljoin
import openpyxl
import csv
from datetime import datetime, timedelta
import traceback
import os
import sys
import base64
import random
import string
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import yaml

# ============ 配置 (支持从环境变量加载) ============
# 支持云端 Github Actions，配置默认留在本地开发用
CSV_PATH = os.environ.get("CSV_PATH", "/Users/apple/Downloads/shell/airports.csv")

TEST_LIMIT = int(os.environ.get("TEST_LIMIT", 0))
START_FROM = int(os.environ.get("START_FROM", 0))
OUTPUT_EXCEL = os.environ.get("OUTPUT_EXCEL", "/Users/apple/Downloads/shell/注册结果.xlsx")
LOG_FILE = os.environ.get("LOG_FILE", "/Users/apple/Downloads/shell/register_log.txt")
NODES_OUTPUT_FILE = os.environ.get("NODES_OUTPUT_FILE", "/Users/apple/Downloads/shell/all_nodes.txt")
MAX_WORKERS = int(os.environ.get("MAX_WORKERS", 8))
CLASH_OUTPUT_FILE = os.environ.get("CLASH_OUTPUT_FILE", "/Users/apple/Downloads/shell/clash_config.yaml")

# 线程锁，确保多线程下日志和数据写入不乱序
stats_lock = threading.Lock()
log_lock = threading.Lock()

REG_EMAIL = os.environ.get("REG_EMAIL", "moneyflysubssr@gmail.com")
REG_PASSWORD = os.environ.get("REG_PASSWORD", "Sikeming001@")

IMAP_SERVER = os.environ.get("IMAP_SERVER", "imap.gmail.com")
IMAP_PORT = int(os.environ.get("IMAP_PORT", 993))
IMAP_EMAIL = os.environ.get("IMAP_EMAIL", "moneyflysubssr@gmail.com")
IMAP_PASSWORD = os.environ.get("IMAP_PASSWORD", "yjqebywkjiokxarx")

TIMEOUT = 10
FAST_TIMEOUT = 5
DELAY_BETWEEN_SITES = 1

BASE_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Content-Type": "application/json",
}

# 临时邮箱服务列表 (按优先级排序，轮询+自动降级)
TEMP_EMAIL_SERVICES = ["mailtm", "tempmail_lol", "guerrillamail"]
# 记录失败的服务，避免反复重试
_FAILED_SERVICES = set()


# ============ 日志 ============
def log(msg):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    formatted_msg = f"[{now}] {msg}"
    with log_lock:
        print(formatted_msg)
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(formatted_msg + "\n")

# ============ 工具函数 ============
def extract_base_url(url):
    parsed = urlparse(url)
    return f"{parsed.scheme}://{parsed.netloc}"


def extract_invite_code(url):
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    invite_keys = [
        "code", "aff", "invite", "invite_code", "invcode", "invitecode",
        "invite-code", "PartnerCode", "affid", "affiliate", "affiliate_code",
        "ref", "referrer", "r", "c",
    ]
    for key in invite_keys:
        if key in params:
            return params[key][0]
    if parsed.fragment:
        frag_str = parsed.fragment
        if "?" in frag_str:
            frag_query = frag_str.split("?", 1)[1]
            frag_params = parse_qs(frag_query)
            hash_keys = invite_keys + ["inviteCode"]
            for key in hash_keys:
                if key in frag_params:
                    return frag_params[key][0]
    return ""


def get_headers(base_url):
    headers = BASE_HEADERS.copy()
    headers["Origin"] = base_url
    headers["Referer"] = base_url + "/"
    return headers


def create_session():
    """创建带 cloudscraper 的 session，可绕过基础 CF 防护"""
    try:
        session = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "darwin", "mobile": False}
        )
    except Exception:
        session = requests.Session()
    retries = Retry(total=1, backoff_factor=0.5, status_forcelist=[502, 503, 504])
    session.mount("http://", HTTPAdapter(max_retries=retries))
    session.mount("https://", HTTPAdapter(max_retries=retries))
    return session


def quick_connectivity_check(base_url, session):
    """快速检测站点是否可达，避免对不可达站点遍历所有端点"""
    try:
        resp = session.get(base_url, timeout=FAST_TIMEOUT, allow_redirects=True)
        return resp.status_code < 500
    except Exception:
        return False


def is_cloudflare_blocked(response):
    if response.status_code in [403, 503]:
        text = response.text[:1000].lower()
        if any(kw in text for kw in ["cloudflare", "cf-", "just a moment", "challenge", "turnstile", "cf-ray"]):
            return True
        content_type = response.headers.get("content-type", "")
        if "text/html" in content_type and len(response.text) > 1000:
            if "<title>" in text and ("attention" in text or "moment" in text):
                return True
    return False


def is_json_response(response):
    content_type = response.headers.get("content-type", "")
    return "application/json" in content_type


def detect_panel_type(base_url, session, headers):
    """探测面板类型: v2board/xboard 还是 sspanel，避免盲目遍历"""
    try:
        resp = session.get(base_url, headers=headers, timeout=FAST_TIMEOUT, allow_redirects=True)
        if is_cloudflare_blocked(resp):
            return "cloudflare"
        text = resp.text[:8000].lower()
        # SSPanel 特征 — 需要更精确匹配 (服务端渲染的表单页面)
        if any(kw in text for kw in ["sspanel", "sspanel-uim"]):
            return "sspanel"
        # SSPanel 通常有服务端渲染的 form action="/auth/login"
        if 'action="/auth/login"' in text or 'action="/auth/register"' in text:
            return "sspanel"
        # V2Board / XBoard SPA 特征
        if any(kw in text for kw in ["v2board", "xboard", "/api/v1/"]):
            return "v2board"
        # 通用 SPA (Vue/React) — 大概率是 V2Board/XBoard
        if '<div id="app"' in text or "app.js" in text or "chunk-vendors" in text or "manifest.js" in text:
            return "v2board"
        # 如果页面有 /auth/ 路径引用但不是 form action，可能是 SPA 路由
        if "#/auth" in text or "/#/login" in text:
            return "v2board"
    except Exception:
        pass
    return "unknown"


def detect_api_base_from_js(base_url, session, headers):
    """从 SPA 的 JS bundle 中提取真实 API 基地址"""
    try:
        resp = session.get(base_url, headers=headers, timeout=FAST_TIMEOUT)
        if resp.status_code != 200:
            return None
        js_urls = re.findall(r'(?:src|href)=["\']([^"\']*\.js[^"\']*)["\']', resp.text)
        priority = [u for u in js_urls if any(kw in u.lower() for kw in ["app.", "main.", "index.", "chunk-"])]
        js_urls = (priority + js_urls)[:5]
        for js_path in js_urls:
            js_url = urljoin(base_url, js_path)
            try:
                js_resp = session.get(js_url, headers=headers, timeout=FAST_TIMEOUT)
                if js_resp.status_code != 200:
                    continue
                js_text = js_resp.text[:50000]
                api_patterns = [
                    r'baseURL\s*[:=]\s*["\']+(https?://[^"\']+?)["\']',
                    r'VUE_APP_API\s*[:=]\s*["\']+(https?://[^"\']+?)["\']',
                    r'apiUrl\s*[:=]\s*["\']+(https?://[^"\']+?)["\']',
                    r'API_URL\s*[:=]\s*["\']+(https?://[^"\']+?)["\']',
                    r'["\']+(https?://[^"\']*?/api/v1)["\']',
                ]
                for pattern in api_patterns:
                    matches = re.findall(pattern, js_text)
                    for match in matches:
                        api_base = match.rstrip("/")
                        if api_base != base_url.rstrip("/") and len(api_base) > 10:
                            log(f"  从JS发现API地址: {api_base}")
                            return api_base
            except Exception:
                pass
    except Exception:
        pass
    return None


def detect_api_base_from_env_js(base_url, session, headers):
    """从 /env.js 中提取 routerBase（真实 API 地址）"""
    try:
        env_url = base_url + "/env.js"
        resp = session.get(env_url, headers=headers, timeout=FAST_TIMEOUT)
        if resp.status_code != 200:
            return None
        match = re.search(r'window\.routerBase\s*=\s*["\']+(https?://[^"\']+)["\']', resp.text)
        if match:
            api_base = match.group(1).rstrip("/")
            if api_base != base_url.rstrip("/"):
                log(f"  从env.js发现API: {api_base}")
                return api_base
    except Exception:
        pass
    return None


def probe_v2board_config(base_url, session, headers):
    """探测 /api/v1/guest/comm/config — V2Board 公开配置端点"""
    try:
        url = base_url + "/api/v1/guest/comm/config"
        resp = session.get(url, headers=headers, timeout=FAST_TIMEOUT)
        if resp.status_code == 200 and is_json_response(resp):
            data = resp.json()
            if "data" in data and isinstance(data["data"], dict):
                config = data["data"]
                log(f"  V2Board 配置: 需邮箱验证={config.get('is_email_verify')}, 需邀请码={config.get('is_invite_force')}")
                return True, config
    except Exception:
        pass
    return False, None



# ============ 临时邮箱多服务轮询系统 ============
# 支持 7 个服务自动轮询 + Gmail IMAP 兜底
# 服务失败自动标记，下次跳过

class TempEmailManager:
    """临时邮箱管理器 — 7 个服务轮询 + 自动降级"""

    # 按优先级排序的服务列表
    SERVICES = [
        "tempmail_lol",      # 极简 API，最稳定
        "mail_tm",           # REST + JWT，域名丰富
        "temp_mail_io",      # REST，速度快
        "throwawaymail_app", # 专为自动化设计
        "maildrop_cc",       # GraphQL，无需注册
        "guerrillamail",     # 老牌稳定
        "dropmail_me",       # GraphQL，多域名
    ]

    def __init__(self):
        self._failed = set()  # 失败的服务
        self._current = None  # 当前使用的服务信息

    def mark_failed(self, service_name):
        self._failed.add(service_name)
        log(f"  ⚠️ 临时邮箱服务 {service_name} 标记为不可用")

    def create_email(self):
        """尝试创建临时邮箱，自动轮询所有服务"""
        for svc in self.SERVICES:
            if svc in self._failed:
                continue
            try:
                result = getattr(self, f"_create_{svc}")()
                if result and result.get("email"):
                    result["service"] = svc
                    self._current = result
                    log(f"  📧 临时邮箱: {result['email']} (via {svc})")
                    return result
            except Exception as e:
                log(f"  {svc} 创建失败: {str(e)[:60]}")
                self.mark_failed(svc)
        log("  ⚠️ 所有临时邮箱服务不可用，使用 Gmail")
        return {"email": REG_EMAIL, "service": "gmail"}

    def fetch_code(self, account_info, wait_seconds=60, check_interval=5):
        """从临时邮箱获取验证码"""
        svc = account_info.get("service", "")
        if svc == "gmail":
            domain = account_info.get("site_domain", "")
            return fetch_verification_code(domain, wait_seconds=wait_seconds,
                                           sent_after=account_info.get("sent_after"))
        fetch_fn = getattr(self, f"_fetch_{svc}", None)
        if not fetch_fn:
            return None
        return fetch_fn(account_info, wait_seconds, check_interval)

    # ---- 1. tempmail.lol ----
    def _create_tempmail_lol(self):
        r = requests.get("https://api.tempmail.lol/v2/inbox/create", timeout=8)
        if r.status_code in (200, 201):
            d = r.json()
            return {"email": d["address"], "token": d["token"]}
        return None

    def _fetch_tempmail_lol(self, info, wait_seconds=60, check_interval=5):
        token = info.get("token")
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                r = requests.get(f"https://api.tempmail.lol/v2/inbox?token={token}", timeout=8)
                if r.status_code == 200:
                    emails = r.json().get("emails", [])
                    for m in emails:
                        body = m.get("body", "") + m.get("html", "") + m.get("subject", "")
                        code = self._extract_code(body)
                        if code:
                            return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 2. mail.tm ----
    def _create_mail_tm(self):
        h = {"Accept": "application/ld+json", "Content-Type": "application/json"}
        r = requests.get("https://api.mail.tm/domains?page=1", headers=h, timeout=8)
        domains = [d["domain"] for d in r.json().get("hydra:member", [])]
        if not domains:
            return None
        username = "".join(random.choices(string.ascii_lowercase + string.digits, k=10))
        address = f"{username}@{domains[0]}"
        password = "Tmp" + "".join(random.choices(string.ascii_letters + string.digits, k=10)) + "!"
        r2 = requests.post("https://api.mail.tm/accounts",
                           json={"address": address, "password": password}, headers=h, timeout=8)
        if r2.status_code != 201:
            return None
        r3 = requests.post("https://api.mail.tm/token",
                           json={"address": address, "password": password}, headers=h, timeout=8)
        if r3.status_code != 200:
            return None
        return {"email": address, "token": r3.json()["token"], "account_id": r2.json().get("id")}

    def _fetch_mail_tm(self, info, wait_seconds=60, check_interval=5):
        token = info.get("token")
        h = {"Authorization": f"Bearer {token}", "Accept": "application/ld+json"}
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                r = requests.get("https://api.mail.tm/messages?page=1", headers=h, timeout=8)
                if r.status_code == 200:
                    msgs = r.json().get("hydra:member", [])
                    for m in msgs:
                        msg_id = m["id"]
                        r2 = requests.get(f"https://api.mail.tm/messages/{msg_id}", headers=h, timeout=8)
                        if r2.status_code == 200:
                            d = r2.json()
                            body = (d.get("text", "") or "") + (d.get("html", [{}])[0] if isinstance(d.get("html"), list) else str(d.get("html", "")))
                            body += d.get("subject", "")
                            code = self._extract_code(body)
                            if code:
                                return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 3. temp-mail.io ----
    def _create_temp_mail_io(self):
        h = {"Accept": "application/json", "Content-Type": "application/json",
             "User-Agent": "Mozilla/5.0"}
        r = requests.post("https://api.internal.temp-mail.io/api/v3/email/new",
                          json={"min_name_length": 10, "max_name_length": 10},
                          headers=h, timeout=8)
        if r.status_code == 200:
            d = r.json()
            return {"email": d["email"], "token": d.get("token", "")}
        return None

    def _fetch_temp_mail_io(self, info, wait_seconds=60, check_interval=5):
        addr = info.get("email")
        h = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                r = requests.get(f"https://api.internal.temp-mail.io/api/v3/email/{addr}/messages",
                                 headers=h, timeout=8)
                if r.status_code == 200:
                    msgs = r.json() if isinstance(r.json(), list) else []
                    for m in msgs:
                        body = m.get("body_text", "") + m.get("body_html", "") + m.get("subject", "")
                        code = self._extract_code(body)
                        if code:
                            return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 4. throwawaymail.app ----
    def _create_throwawaymail_app(self):
        h = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
        r = requests.post("https://throwawaymail.app/api/mailboxes", headers=h, timeout=8)
        if r.status_code == 201:
            d = r.json()
            return {"email": d["address"], "mailbox_id": d["mailbox_id"]}
        return None

    def _fetch_throwawaymail_app(self, info, wait_seconds=60, check_interval=5):
        mid = info.get("mailbox_id")
        h = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                r = requests.get(f"https://throwawaymail.app/api/mailboxes/{mid}/messages",
                                 headers=h, timeout=8)
                if r.status_code == 200:
                    msgs = r.json() if isinstance(r.json(), list) else []
                    for m in msgs:
                        # 获取完整内容
                        msg_id = m.get("id", "")
                        if msg_id:
                            r2 = requests.get(f"https://throwawaymail.app/api/mailboxes/{mid}/messages/{msg_id}",
                                              headers=h, timeout=8)
                            if r2.status_code == 200:
                                d = r2.json()
                                body = d.get("body", "") + d.get("subject", "")
                                code = self._extract_code(body)
                                if code:
                                    return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 5. maildrop.cc ----
    def _create_maildrop_cc(self):
        username = "reg" + "".join(random.choices(string.ascii_lowercase + string.digits, k=8))
        return {"email": f"{username}@maildrop.cc", "mailbox": username}

    def _fetch_maildrop_cc(self, info, wait_seconds=60, check_interval=5):
        mailbox = info.get("mailbox")
        h = {"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                q = f'{{"query":"query {{ inbox(mailbox:\\"{mailbox}\\") {{ id subject headerfrom date html }} }}"}}'
                r = requests.post("https://api.maildrop.cc/graphql", data=q, headers=h, timeout=8)
                if r.status_code == 200:
                    msgs = r.json().get("data", {}).get("inbox", [])
                    for m in msgs:
                        body = (m.get("html", "") or "") + (m.get("subject", "") or "")
                        code = self._extract_code(body)
                        if code:
                            return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 6. guerrillamail ----
    def _create_guerrillamail(self):
        r = requests.get("https://api.guerrillamail.com/ajax.php?f=get_email_address&lang=en",
                          timeout=8)
        if r.status_code == 200:
            d = r.json()
            return {"email": d["email_addr"], "sid_token": d["sid_token"]}
        return None

    def _fetch_guerrillamail(self, info, wait_seconds=60, check_interval=5):
        sid = info.get("sid_token")
        end = time.time() + wait_seconds
        seq = 0
        time.sleep(5)
        while time.time() < end:
            try:
                r = requests.get(
                    f"https://api.guerrillamail.com/ajax.php?f=check_email&seq={seq}&sid_token={sid}",
                    timeout=8)
                if r.status_code == 200:
                    d = r.json()
                    msgs = d.get("list", [])
                    for m in msgs:
                        if m.get("mail_from", "") == "no-reply@guerrillamail.com":
                            continue  # 跳过欢迎邮件
                        mail_id = m.get("mail_id")
                        r2 = requests.get(
                            f"https://api.guerrillamail.com/ajax.php?f=fetch_email&email_id={mail_id}&sid_token={sid}",
                            timeout=8)
                        if r2.status_code == 200:
                            md = r2.json()
                            body = md.get("mail_body", "") + md.get("mail_subject", "")
                            code = self._extract_code(body)
                            if code:
                                return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 7. dropmail.me ----
    def _create_dropmail_me(self):
        q = "mutation{introduceSession{id,expiresAt,addresses{address}}}"
        r = requests.post("https://dropmail.me/api/graphql/web-test-2",
                          json={"query": q}, timeout=8,
                          headers={"User-Agent": "Mozilla/5.0"})
        if r.status_code == 200:
            d = r.json().get("data", {}).get("introduceSession", {})
            addrs = d.get("addresses", [])
            if addrs:
                return {"email": addrs[0]["address"], "session_id": d["id"]}
        return None

    def _fetch_dropmail_me(self, info, wait_seconds=60, check_interval=5):
        sid = info.get("session_id")
        end = time.time() + wait_seconds
        time.sleep(5)
        while time.time() < end:
            try:
                q = f'{{session(id:"{sid}"){{mails{{rawSize,headerSubject,text,headerFrom}}}}}}'
                r = requests.post("https://dropmail.me/api/graphql/web-test-2",
                                  json={"query": q}, timeout=8,
                                  headers={"User-Agent": "Mozilla/5.0"})
                if r.status_code == 200:
                    mails = r.json().get("data", {}).get("session", {}).get("mails", [])
                    for m in mails:
                        body = (m.get("text", "") or "") + (m.get("headerSubject", "") or "")
                        code = self._extract_code(body)
                        if code:
                            return code
            except Exception:
                pass
            time.sleep(check_interval)
        return None

    # ---- 通用验证码提取 ----
    @staticmethod
    def _extract_code(text):
        """从邮件正文中提取 4-6 位验证码"""
        if not text:
            return None
        # 优先匹配"验证码"关键词旁的数字
        patterns = [
            r'验证码[：:\s]*(\d{4,6})',
            r'code[:\s]*(\d{4,6})',
            r'Code[:\s]*(\d{4,6})',
            r'<b>(\d{4,6})</b>',
            r'<strong>(\d{4,6})</strong>',
            r'\b(\d{6})\b',
            r'\b(\d{4,5})\b',
        ]
        for p in patterns:
            m = re.search(p, text, re.I)
            if m:
                return m.group(1)
        return None


# 全局单例
temp_email_mgr = TempEmailManager()


def get_registration_email(session, config):
    """根据站点配置选择邮箱: 优先临时邮箱，必要时使用 Gmail"""
    whitelist = (config or {}).get("email_whitelist_suffix", "")
    if whitelist:
        suffixes = [s.strip().lower() for s in str(whitelist).split(",") if s.strip()]
        gmail_ok = any(s in ("gmail.com", "@gmail.com") for s in suffixes)
        if suffixes and not gmail_ok:
            log(f"  站点限制邮箱后缀: {suffixes}，使用 Gmail")
            return {"email": REG_EMAIL, "service": "gmail"}

    # 使用临时邮箱轮询系统
    result = temp_email_mgr.create_email()
    return result



# ============ 导航页处理 ============
def extract_actual_url(url, session):
    if "guatizi.com" not in url:
        return url
    log(f"  检测到导航页，尝试提取真实 URL...")
    try:
        resp = session.get(url, headers=BASE_HEADERS, timeout=TIMEOUT)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.text, "html.parser")
            go_links = soup.find_all("a", href=re.compile(r"/go/\?url="))
            for link in go_links:
                href = link.get("href")
                parsed_href = urlparse(href)
                query = parse_qs(parsed_href.query)
                if "url" in query:
                    encoded_url = query["url"][0]
                    try:
                        decoded_url = base64.b64decode(encoded_url).decode("utf-8")
                        if "guatizi.com" in decoded_url:
                            log(f"  二级跳转: {decoded_url}")
                            resp2 = session.get(decoded_url, headers=BASE_HEADERS, timeout=TIMEOUT, allow_redirects=True)
                            if resp2.status_code == 200:
                                return extract_base_url(resp2.url)
                        return extract_base_url(decoded_url)
                    except Exception:
                        pass
            text_urls = re.findall(r'https?://[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', resp.text)
            for t_url in text_urls:
                if "guatizi.com" not in t_url and "github.com" not in t_url and "t.me" not in t_url:
                    return extract_base_url(t_url)
    except Exception as e:
        log(f"  提取真实 URL 失败: {e}")
    return url


# ============ IMAP 验证码获取 (优化版) ============
def extract_main_domain(domain):
    parts = domain.lower().split(".")
    if len(parts) >= 2:
        return ".".join(parts[-2:])
    return domain.lower()


def fetch_verification_code(site_domain, wait_seconds=60, check_interval=5, sent_after=None):
    """从 IMAP 获取验证码 — 只取 sent_after 之后的最新含验证码邮件"""
    main_domain = extract_main_domain(site_domain)
    log(f"  等待验证码邮件 (匹配域名: {main_domain})...")
    start_time = time.time()
    if sent_after is None:
        sent_after = datetime.now() - timedelta(seconds=10)

    folders = ["INBOX", "[Gmail]/Spam", "[Gmail]/All Mail", "[Gmail]/Trash"]

    time.sleep(5)

    while time.time() - start_time < wait_seconds:
        try:
            mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
            mail.login(IMAP_EMAIL, IMAP_PASSWORD)

            for folder in folders:
                try:
                    status, _ = mail.select(folder)
                    if status != "OK":
                        continue
                except Exception:
                    continue

                since_date = (datetime.now() - timedelta(minutes=5)).strftime("%d-%b-%Y")
                status, messages = mail.search(None, f'(SINCE {since_date})')

                if status == "OK" and messages[0]:
                    mail_ids = messages[0].split()
                    # 只看最新的几封邮件
                    for mid in reversed(mail_ids[-10:]):
                        status, msg_data = mail.fetch(mid, "(RFC822)")
                        if status != "OK":
                            continue
                        msg = email.message_from_bytes(msg_data[0][1])

                        # 严格时间过滤：只要 sent_after 之后的邮件
                        date_str = msg.get("Date", "")
                        try:
                            msg_date = email.utils.parsedate_to_datetime(date_str)
                            if msg_date.tzinfo:
                                utc_offset = datetime.now().astimezone().utcoffset()
                                msg_date_local = msg_date.replace(tzinfo=None) + (utc_offset or timedelta(0))
                            else:
                                msg_date_local = msg_date
                            if msg_date_local < sent_after.replace(tzinfo=None):
                                continue
                        except Exception:
                            continue  # 无法解析时间的邮件直接跳过

                        # 提取邮件正文
                        body = ""
                        if msg.is_multipart():
                            for part in msg.walk():
                                ct = part.get_content_type()
                                if ct in ("text/plain", "text/html"):
                                    try:
                                        body += part.get_payload(decode=True).decode("utf-8", errors="replace")
                                    except Exception:
                                        pass
                        else:
                            try:
                                body = msg.get_payload(decode=True).decode("utf-8", errors="replace")
                            except Exception:
                                pass

                        subject = ""
                        try:
                            subj_raw, enc = decode_header(msg["Subject"])[0]
                            if isinstance(subj_raw, bytes):
                                subject = subj_raw.decode(enc or "utf-8", errors="replace")
                            else:
                                subject = str(subj_raw)
                        except Exception:
                            pass

                        from_addr = msg.get("From", "").lower()

                        # 从正文+标题中提取验证码
                        all_text = subject + " " + body
                        codes = re.findall(r'\b(\d{4,6})\b', all_text)
                        if codes:
                            code = codes[0]
                            log(f"  找到验证码: {code} (来自: {from_addr[:50]})")
                            mail.logout()
                            return code

            mail.logout()
        except Exception as e:
            log(f"  IMAP 错误: {e}")

        time.sleep(check_interval)

    log(f"  未能在 {wait_seconds}s 内获取验证码")
    return None

# ============ 核心注册/登录逻辑 ============
def get_endpoints_for_panel(panel_type, action):
    """根据面板类型返回优先端点列表，避免盲目遍历"""
    if action == "login":
        if panel_type == "v2board":
            return ["/api/v1/passport/auth/login"]
        elif panel_type == "sspanel":
            return ["/auth/login"]
        else:
            return [
                "/api/v1/passport/auth/login",
                "/auth/login",
                "/api/v2/passport/auth/login",
            ]
    elif action == "register":
        if panel_type == "v2board":
            return ["/api/v1/passport/auth/register"]
        elif panel_type == "sspanel":
            return ["/auth/register"]
        else:
            return [
                "/api/v1/passport/auth/register",
                "/auth/register",
                "/api/v2/passport/auth/register",
            ]
    elif action == "send_code":
        if panel_type == "sspanel":
            return ["/auth/send", "/api/auth/sendEmailVerify"]
        else:
            return [
                "/api/v1/passport/comm/sendEmailVerify",
                "/api/comm/sendEmailVerify",
            ]
    return []


def try_api_login(base_url, headers, session, email, panel_type="unknown"):
    endpoints = get_endpoints_for_panel(panel_type, "login")
    for endpoint in endpoints:
        url = base_url + endpoint
        try:
            resp = session.post(url, json={
                "email": email,
                "password": REG_PASSWORD
            }, headers=headers, timeout=TIMEOUT)

            if is_cloudflare_blocked(resp):
                return {"status": "cloudflare", "endpoint": endpoint}

            if is_json_response(resp):
                data = resp.json()
                if "data" in data and data.get("data"):
                    token = None
                    if isinstance(data["data"], dict):
                        token = data["data"].get("auth_data") or data["data"].get("token")
                    return {"status": "login_ok", "data": data, "token": token, "endpoint": endpoint}
                if data.get("ret") == 1:
                    token = data.get("token") or (data.get("data", {}).get("token") if isinstance(data.get("data"), dict) else None)
                    return {"status": "login_ok", "data": data, "token": token, "endpoint": endpoint}
                log(f"  登录返回: {endpoint} -> {json.dumps(data, ensure_ascii=False)[:200]}")
                return {"status": "login_fail", "data": data, "endpoint": endpoint}
            else:
                if resp.status_code in [403, 503]:
                    return {"status": "cloudflare", "endpoint": endpoint}
        except requests.exceptions.Timeout:
            log(f"  登录超时: {endpoint}")
        except requests.exceptions.ConnectionError:
            pass
        except Exception as e:
            log(f"  登录异常: {endpoint} - {e}")
    return {"status": "no_api"}

def try_api_register(base_url, headers, session, email, invite_code="", panel_type="unknown"):
    endpoints = get_endpoints_for_panel(panel_type, "register")
    for endpoint in endpoints:
        url = base_url + endpoint
        try:
            payload = {
                "email": email,
                "password": REG_PASSWORD,
                "passwd": REG_PASSWORD,
                "repasswd": REG_PASSWORD,
                "password_confirmation": REG_PASSWORD,
            }
            if invite_code:
                payload["invite_code"] = invite_code
                payload["code"] = invite_code
                payload["aff"] = invite_code
            # SSPanel 需要同意服务条款
            if panel_type == "sspanel" or "/auth/" in endpoint:
                payload["agree"] = 1
                payload["tos"] = 1
                payload["name"] = email.split("@")[0]
                payload["name_again"] = email.split("@")[0]
                payload["email_code"] = ""

            resp = session.post(url, json=payload, headers=headers, timeout=TIMEOUT)

            if is_cloudflare_blocked(resp):
                return {"status": "cloudflare", "endpoint": endpoint}

            if is_json_response(resp):
                data = resp.json()
                msg = str(data.get("message", data.get("msg", ""))).lower()
                errors = str(data.get("errors", "")).lower()
                combined = msg + " " + errors

                if any(kw in combined for kw in ["后缀", "白名单", "黑名单", "临时邮箱", "不允许", "禁止", "不合法", "invalid email", "not allowed", "reject", "permit"]):
                    return {"status": "email_rejected", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["已被使用", "already", "已注册", "exist", "registered", "邮箱已存在"]):
                    return {"status": "already_registered", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["图形验证码", "captcha", "geetest", "turnstile", "recaptcha", "人机验证"]):
                    return {"status": "need_captcha", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["验证码", "email_code", "verify", "verification", "邮箱验证"]):
                    return {"status": "need_email_code", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["邀请码", "invite", "invitation"]):
                    return {"status": "need_invite", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["服务条款", "tos", "agree"]):
                    return {"status": "need_tos", "data": data, "endpoint": endpoint}

                if any(kw in combined for kw in ["关闭注册", "停止注册", "暂停注册", "registration closed", "not open"]):
                    return {"status": "closed", "data": data, "endpoint": endpoint}

                if "data" in data and data.get("data"):
                    token = None
                    if isinstance(data["data"], dict):
                        token = data["data"].get("auth_data") or data["data"].get("token")
                    return {"status": "register_ok", "data": data, "token": token, "endpoint": endpoint}

                if data.get("status") == True or data.get("ret") == 1:
                    token = None
                    if isinstance(data.get("data"), dict):
                        token = data["data"].get("auth_data") or data["data"].get("token")
                    return {"status": "register_ok", "data": data, "token": token, "endpoint": endpoint}

                log(f"  注册返回: {endpoint} -> {json.dumps(data, ensure_ascii=False)[:200]}")
                return {"status": "register_fail", "data": data, "endpoint": endpoint}
            else:
                if resp.status_code in [403, 503]:
                    return {"status": "cloudflare", "endpoint": endpoint}
                log(f"  注册非JSON: {endpoint} -> status={resp.status_code}")
        except requests.exceptions.Timeout:
            log(f"  注册超时: {endpoint}")
        except requests.exceptions.ConnectionError:
            pass
        except Exception as e:
            log(f"  注册异常: {endpoint} - {e}")
    return {"status": "no_api"}

def try_register_with_email_code(base_url, headers, session, email, invite_code="", endpoint="", panel_type="unknown", email_info=None):
    """需要邮件验证码的注册流程"""
    domain = urlparse(base_url).netloc
    send_endpoints = get_endpoints_for_panel(panel_type, "send_code")

    code_sent = False
    send_time = datetime.now()
    for ep in send_endpoints:
        url = base_url + ep
        try:
            resp = session.post(url, json={"email": email}, headers=headers, timeout=TIMEOUT)
            if is_cloudflare_blocked(resp):
                return {"status": "cloudflare", "endpoint": ep}
            if is_json_response(resp):
                data = resp.json()
                msg = str(data.get("message", data.get("msg", ""))).lower()
                errors = str(data.get("errors", "")).lower()
                combined = msg + " " + errors
                if any(kw in combined for kw in ["后缀", "白名单", "黑名单", "临时邮箱", "不允许", "禁止", "不合法", "invalid email", "not allowed", "reject", "permit"]):
                    code_sent = False
                    return {"status": "email_rejected", "data": data}

                if any(kw in combined for kw in ["图形验证码", "captcha", "geetest", "turnstile", "recaptcha", "人机验证"]):
                    code_sent = False
                    return {"status": "need_captcha", "data": data}

                if data.get("data") is not None or "success" in str(data).lower() or data.get("ret") == 1:
                    log(f"  验证码已发送到 {email}")
                    code_sent = True
                    break
                else:
                    log(f"  发送验证码响应: {json.dumps(data, ensure_ascii=False)[:150]}")
        except Exception as e:
            log(f"  发送验证码失败: {ep} - {e}")

    if not code_sent:
        return {"status": "send_code_fail"}

    # 根据邮箱类型选择验证码获取方式
    if email_info and email_info.get("service") not in ("gmail", None):
        # 临时邮箱 — 通过 TempEmailManager 轮询收信
        code = temp_email_mgr.fetch_code(email_info, wait_seconds=60)
    else:
        # Gmail IMAP
        code = fetch_verification_code(domain, wait_seconds=60, sent_after=send_time)
    if not code:
        return {"status": "no_code_received"}

    if not endpoint:
        endpoints = get_endpoints_for_panel(panel_type, "register")
        endpoint = endpoints[0] if endpoints else "/api/v1/passport/auth/register"

    url = base_url + endpoint
    try:
        payload = {
            "email": email,
            "password": REG_PASSWORD,
            "passwd": REG_PASSWORD,
            "repasswd": REG_PASSWORD,
            "password_confirmation": REG_PASSWORD,
            "email_code": code,
        }
        if invite_code:
            payload["invite_code"] = invite_code
            payload["code"] = invite_code
            payload["aff"] = invite_code
        if panel_type == "sspanel" or "/auth/" in endpoint:
            payload["agree"] = 1
            payload["tos"] = 1
            payload["name"] = email.split("@")[0]

        resp = session.post(url, json=payload, headers=headers, timeout=TIMEOUT)
        if is_json_response(resp):
            data = resp.json()
            if "data" in data and data.get("data"):
                token = None
                if isinstance(data["data"], dict):
                    token = data["data"].get("auth_data") or data["data"].get("token")
                return {"status": "register_ok", "data": data, "token": token}
            if data.get("ret") == 1 or data.get("status") == True:
                token = None
                if isinstance(data.get("data"), dict):
                    token = data["data"].get("auth_data") or data["data"].get("token")
                return {"status": "register_ok", "data": data, "token": token}
            log(f"  验证码注册返回: {json.dumps(data, ensure_ascii=False)[:200]}")
            return {"status": "register_fail", "data": data}
    except Exception as e:
        log(f"  验证码注册异常: {e}")
    return {"status": "register_fail"}


def try_register_with_tos_retry(base_url, headers, session, email, invite_code="", endpoint="", panel_type="unknown"):
    """SSPanel 需要同意服务条款的重试"""
    if not endpoint:
        endpoint = "/auth/register"
    url = base_url + endpoint
    try:
        payload = {
            "email": email,
            "password": REG_PASSWORD,
            "passwd": REG_PASSWORD,
            "repasswd": REG_PASSWORD,
            "agree": 1,
            "tos": 1,
            "name": email.split("@")[0],
        }
        if invite_code:
            payload["invite_code"] = invite_code
            payload["code"] = invite_code
        resp = session.post(url, json=payload, headers=headers, timeout=TIMEOUT)
        if is_json_response(resp):
            data = resp.json()
            if data.get("ret") == 1 or ("data" in data and data.get("data")):
                token = None
                if isinstance(data.get("data"), dict):
                    token = data["data"].get("auth_data") or data["data"].get("token")
                return {"status": "register_ok", "data": data, "token": token}
            msg = str(data.get("msg", data.get("message", ""))).lower()
            if "验证码" in msg or "email_code" in msg:
                return {"status": "need_email_code", "data": data, "endpoint": endpoint}
            log(f"  TOS重试返回: {json.dumps(data, ensure_ascii=False)[:200]}")
    except Exception as e:
        log(f"  TOS重试异常: {e}")
    return {"status": "register_fail"}

def fetch_and_save_nodes(sub_url, site_name="未知"):
    """请求订阅链接并提取节点保存到本地"""
    if not sub_url:
        return 0

def convert_clash_proxy_to_uri(proxy):
    """
    将 Clash 格式的代理字典逆向转换为标准 URI 链接
    """
    try:
        p_type = str(proxy.get("type", "")).lower()
        name = urllib.parse.quote(proxy.get("name", "node"))
        server = proxy.get("server", "")
        port = proxy.get("port", "")
        
        if p_type == "vmess":
            # 还原 vmess://
            js = {
                "v": "2", "ps": proxy.get("name"), "add": server, "port": str(port),
                "id": proxy.get("uuid"), "aid": str(proxy.get("alterId", 0)),
                "net": proxy.get("network", "tcp"), "type": "none",
                "host": proxy.get("ws-opts", {}).get("headers", {}).get("Host", ""),
                "path": proxy.get("ws-opts", {}).get("path", ""),
                "tls": "tls" if proxy.get("tls") else ""
            }
            content = base64.b64encode(json.dumps(js).encode('utf-8')).decode('utf-8')
            return f"vmess://{content}"
            
        elif p_type == "vless":
            # 还原 vless://
            uuid = proxy.get("uuid")
            params = []
            if proxy.get("tls"):
                params.append("security=tls")
            if proxy.get("network"):
                params.append(f"type={proxy.get('network')}")
            if proxy.get("ws-opts"):
                params.append(f"path={urllib.parse.quote(proxy['ws-opts'].get('path', '/'))}")
            
            query = "&".join(params)
            return f"vless://{uuid}@{server}:{port}?{query}#{name}"
            
        elif p_type == "ss":
            # 还原 ss://
            method = proxy.get("cipher", "aes-256-gcm")
            password = proxy.get("password", "")
            # SS 许多格式是 ss://BASE64(method:password)@ip:port#name
            user_info = base64.b64encode(f"{method}:{password}".encode('utf-8')).decode('utf-8')
            return f"ss://{user_info}@{server}:{port}#{name}"
            
        elif p_type == "trojan":
            # 还原 trojan://
            password = proxy.get("password", "")
            return f"trojan://{password}@{server}:{port}#{name}"
            
    except:
        pass
    return None

def extract_and_save_nodes(sub_url, site_name):
    if not sub_url:
        return 0
    try:
        # 使用 v2rayN UA 获取 Base64 原数据
        h = {"User-Agent": "v2rayN/6.23"}
        r = requests.get(sub_url, headers=h, timeout=10)
        text = r.text.strip()
        if not text:
            log(f"  获取订阅连接 [{sub_url}] 失败: 服务端返回空内容")
            return 0
            
        nodes = []
        is_yaml = False
        
        # 1. 尝试直接作为 YAML/Clash 配置解析
        try:
            # 排除简单的 Base64，如果是 YAML 通常包含特定的 keywords
            if "proxies:" in text or "proxy-groups:" in text:
                data = yaml.safe_load(text)
                if isinstance(data, dict) and "proxies" in data:
                    for p in data["proxies"]:
                        uri = convert_clash_proxy_to_uri(p)
                        if uri:
                            nodes.append(uri)
                    is_yaml = True
        except:
            pass
            
        if not is_yaml:
            # 2. 尝试 Base64 解码 (传统 V2Ray/SSR 格式)
            try:
                # 预处理：去除可能的换行并补齐 padding
                cleaned_text = text.replace("\n", "").replace("\r", "")
                missing_padding = len(cleaned_text) % 4
                text_padded = cleaned_text + '=' * ((4 - missing_padding) % 4)
                decoded = base64.b64decode(text_padded).decode('utf-8')
                lines = decoded.splitlines()
            except:
                # 3. 或者非 base64 纯文本一行一个 URI
                lines = text.splitlines()
            
            # 提取常见的节点前缀
            supported_schemes = ("vmess://", "vless://", "ss://", "ssr://", "trojan://", "hysteria", "tuic")
            for line in lines:
                line = line.strip()
                if line.startswith(supported_schemes):
                    nodes.append(line)
        
        if nodes:
            with log_lock:  # 使用统一锁防止并发写入乱序
                with open(NODES_OUTPUT_FILE, "a+", encoding="utf-8") as f:
                    # 确保文件末尾有换行后再追加
                    f.seek(0, 2)
                    if f.tell() > 0:
                        f.seek(f.tell() - 1)
                        if f.read(1) != '\n':
                            f.write('\n')
                    
                    f.write(f"# === {site_name} | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
                    f.write("\n".join(nodes) + "\n")
            log(f"  成功提取并保存了 {len(nodes)} 个节点")
            return len(nodes)
        else:
            log(f"  订阅为空或格式不匹配 (内容前缀预览: {text[:30]}...)")
    except Exception as e:
        log(f"  提取节点失败: {str(e)[:50]}")
    return 0


def extract_subscription(base_url, headers, session, token):
    auth_headers = headers.copy()
    if token:
        if token.lower().startswith("bearer"):
            auth_headers["Authorization"] = token
        else:
            auth_headers["Authorization"] = f"Bearer {token}"

    sub_endpoints = [
        "/api/v1/user/getSubscribe",
        "/api/v1/client/subscribe",
        "/api/user/getSubscribe",
        "/api/v1/user/subscribe",
    ]

    if token:
        for endpoint in sub_endpoints:
            url = base_url + endpoint
            try:
                resp = session.get(url, headers=auth_headers, timeout=TIMEOUT)
                if is_json_response(resp):
                    data = resp.json()
                    if isinstance(data.get("data"), dict):
                        sub_url = data["data"].get("subscribe_url") or data["data"].get("subscribe_link")
                        if sub_url:
                            return {"subscribe_url": sub_url, "subscribe_domain": urlparse(sub_url).netloc}
                        sub_token = data["data"].get("token")
                        if sub_token:
                            return {"subscribe_url": f"{base_url}/api/v1/client/subscribe?token={sub_token}",
                                    "subscribe_domain": urlparse(base_url).netloc}
                elif resp.headers.get("subscription-userinfo") or resp.headers.get("content-disposition"):
                    return {"subscribe_url": url, "subscribe_domain": urlparse(base_url).netloc}
            except Exception:
                pass

    user_paths = ["/user", "/dashboard", "/user/profile"]
    for path in user_paths:
        try:
            resp = session.get(base_url + path, headers=headers, timeout=TIMEOUT)
            if resp.status_code == 200:
                patterns = [
                    r'https?://[^\s"\']+?\?token=[a-zA-Z0-9]{16,}',
                    r'https?://[^\s"\']+?/link/[a-zA-Z0-9]{16,}',
                    r'https?://[^\s"\']+?\.subscribe\?[\w=&]+',
                ]
                for pattern in patterns:
                    matches = re.findall(pattern, resp.text)
                    if matches:
                        sub_url = matches[0]
                        return {"subscribe_url": sub_url, "subscribe_domain": urlparse(sub_url).netloc}
        except Exception:
            pass
    return None


PLAN_OUTPUT_EXCEL = "/Users/apple/Downloads/shell/套餐分析.xlsx"

PERIOD_MONTHS = {
    "month_price": 1, "quarter_price": 3, "half_year_price": 6, "year_price": 12,
    "two_year_price": 24, "three_year_price": 36, "onetime_price": 0,
}


def convert_price_to_yuan(price_raw):
    """
    V2Board/XBoard API 返回的价格单位是"分"（整数），需要除以 100 转为元。
    例如: 9600 -> 96.00 元, 6000 -> 60.00 元
    部分面板可能返回的已经是元（带小数点），需要智能判断。
    """
    if price_raw is None:
        return None
    try:
        val = float(price_raw)
    except (ValueError, TypeError):
        return None
    if val <= 0:
        return None
    # 判断逻辑：如果值是整数且 >= 100，大概率是分（V2Board 标准行为）
    # 如果值带小数点（如 9.9），则认为已经是元
    if val == int(val) and val >= 100:
        return round(val / 100, 2)
    elif val == int(val) and val > 0:
        # 小整数（如 1-99）：可能是 1-99 分 = 0.01-0.99 元，也可能就是 1-99 元
        # V2Board 通常不会设低于 1 元的套餐，所以 1-99 更可能是元
        # 但如果是 1-9 这种极低值，有可能是分（0.01-0.09 元）或元
        # 保守处理：10 以下当元，10-99 之间也当元（常见定价如 10元、29元、59元）
        return val
    else:
        # 带小数点的值，直接当元
        return round(val, 2)


def parse_traffic_gb(transfer_enable):
    """
    解析流量字段，智能判断单位：
    - V2Board/XBoard：transfer_enable 直接就是 GB 数值（如 250 = 250GB）
    - 部分旧面板：可能返回字节数（如 268435456000 = 250GB）
    """
    if not transfer_enable:
        return 0
    try:
        val = float(transfer_enable)
        if val <= 0:
            return 0
        # 超过 10000 认为是字节单位（没有面板会给 10TB+ 的 GB 值）
        if val > 10000:
            # 字节 -> GB
            gb = val / (1024 ** 3)
            if gb < 0.1:
                # 可能是 KB 或 MB 单位
                gb_from_mb = val / 1024
                if gb_from_mb > 0.1:
                    return round(gb_from_mb, 1)
                gb_from_kb = val / (1024 * 1024)
                if gb_from_kb > 0.1:
                    return round(gb_from_kb, 1)
            return round(gb, 1)
        # <= 10000 直接当 GB（V2Board 标准行为）
        return round(val, 1)
    except Exception:
        return 0


def fetch_plans(base_url, headers, session, token, panel_type="unknown"):
    auth_headers = headers.copy()
    if token:
        if str(token).lower().startswith("bearer"):
            auth_headers["Authorization"] = token
        else:
            auth_headers["Authorization"] = f"Bearer {token}"

    plans = []
    plan_endpoints = [
        "/api/v1/guest/plan/fetch",
        "/api/v1/user/plan/fetch",
        "/api/v1/plan/fetch",
    ]

    for ep in plan_endpoints:
        try:
            h = auth_headers if "user" in ep else headers
            resp = session.get(base_url + ep, headers=h, timeout=TIMEOUT)
            if not is_json_response(resp):
                continue
            data = resp.json()
            raw = data.get("data") if isinstance(data.get("data"), list) else (data if isinstance(data, list) else None)
            if not raw:
                continue
            for p in raw:
                if not isinstance(p, dict):
                    continue
                # 跳过不可售卖的套餐
                if p.get("sell") is False or p.get("show") is False:
                    continue
                name = p.get("name", "未知套餐")
                traffic_gb = parse_traffic_gb(p.get("transfer_enable", 0))
                device_limit = p.get("device_limit") or 0
                speed_limit = p.get("speed_limit") or 0  # Mbps
                for price_key, months in PERIOD_MONTHS.items():
                    if months == 0:
                        # onetime_price：一次性付费，不计算月价
                        onetime = convert_price_to_yuan(p.get(price_key))
                        if onetime and onetime > 0:
                            plans.append({
                                "name": name, "price": onetime, "period": "一次性",
                                "months": 0, "monthly_price": 0,
                                "traffic_gb": traffic_gb, "device_limit": device_limit,
                                "speed_limit": speed_limit,
                            })
                        continue
                    price_yuan = convert_price_to_yuan(p.get(price_key))
                    if price_yuan is None or price_yuan <= 0:
                        continue
                    monthly = round(price_yuan / months, 2)
                    period_name = {1:"月付",3:"季付",6:"半年付",12:"年付",24:"两年付",36:"三年付"}.get(months, f"{months}月")
                    plans.append({
                        "name": name, "price": price_yuan, "period": period_name,
                        "months": months, "monthly_price": monthly,
                        "traffic_gb": traffic_gb, "device_limit": device_limit,
                        "speed_limit": speed_limit,
                    })
            if plans:
                log(f"  获取到 {len(plans)} 个套餐选项")
                # 打印关键套餐信息以便验证
                for pl in plans[:3]:
                    log(f"    {pl['name']} | {pl['period']} {pl['price']}元 | {pl['traffic_gb']}GB | 设备:{pl['device_limit']} | 带宽:{pl['speed_limit']}M")
                return plans
        except Exception:
            pass

    if panel_type == "sspanel":
        try:
            resp = session.get(base_url + "/user/shop", headers=auth_headers, timeout=TIMEOUT)
            if resp.status_code == 200 and "text/html" in resp.headers.get("content-type", ""):
                soup = BeautifulSoup(resp.text, "html.parser")
                cards = soup.find_all(class_=re.compile(r"card|plan|product|shop"))
                for card in cards:
                    text = card.get_text(" ", strip=True)
                    name_el = card.find(re.compile(r"h[1-6]"))
                    pname = name_el.get_text(strip=True) if name_el else text[:30]
                    price_m = re.search(r'[¥￥$]\s*(\d+\.?\d*)', text)
                    traffic_m = re.search(r'(\d+\.?\d*)\s*(GB|TB|G|T)', text, re.I)
                    if price_m:
                        price = float(price_m.group(1))
                        tg = 0
                        if traffic_m:
                            tg = float(traffic_m.group(1))
                            if traffic_m.group(2).upper().startswith("T"):
                                tg *= 1024
                        plans.append({"name": pname, "price": price, "period": "月付",
                                      "months": 1, "monthly_price": price,
                                      "traffic_gb": tg, "device_limit": 0,
                                      "speed_limit": 0})
                if plans:
                    log(f"  从网页获取到 {len(plans)} 个套餐")
                    return plans
        except Exception:
            pass
    return plans


def generate_plan_analysis(all_plans_data):
    if not all_plans_data:
        log("无套餐数据，跳过生成套餐分析")
        return
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "全部套餐"
    ws1.append(["站点名称", "套餐名", "月价(元)", "流量(GB)", "设备限制", "带宽(Mbps)", "周期", "原价(元)"])
    for item in all_plans_data:
        for plan in item["plans"]:
            ws1.append([item["site_name"], plan["name"], plan["monthly_price"],
                        plan["traffic_gb"], plan["device_limit"] or "不限",
                        plan.get("speed_limit") or "不限",
                        plan["period"], plan["price"]])

    ws2 = wb.create_sheet("推荐套餐")
    ws2.append(["站点名称", "套餐名", "月价(元)", "流量(GB)", "设备限制", "带宽(Mbps)", "周期", "原价(元)", "性价比(GB/元)"])
    flat = []
    for item in all_plans_data:
        for plan in item["plans"]:
            score = round(plan["traffic_gb"] / plan["monthly_price"], 2) if plan["monthly_price"] > 0 else 0
            flat.append({**plan, "site_name": item["site_name"], "score": score})
    flat.sort(key=lambda x: (-x["score"], x["monthly_price"], -(x["device_limit"] == 0)))
    for p in flat:
        ws2.append([p["site_name"], p["name"], p["monthly_price"], p["traffic_gb"],
                     p["device_limit"] or "不限", p.get("speed_limit") or "不限",
                     p["period"], p["price"], p["score"]])

    wb.save(PLAN_OUTPUT_EXCEL)
    log(f"套餐分析已保存到: {PLAN_OUTPUT_EXCEL}")
    log(f"  全部套餐: {sum(len(i['plans']) for i in all_plans_data)} 条, 涉及 {len(all_plans_data)} 个站点")


def generate_clash_config(all_nodes_text):
    """
    将提取的所有节点链接转换为 Clash 配置文件 (增加命名唯一化逻辑)
    """
    proxies = []
    seen_names = set()
    # 先做一次全局去重
    lines = list(set((all_nodes_text or "").split('\n')))
    
    for line in lines:
        line = line.strip()
        if not line or "://" not in line:
            continue
        
        try:
            node = None
            if line.startswith("vmess://"):
                data_b64 = line[8:]
                data_b64 += "=" * ((4 - len(data_b64) % 4) % 4)
                data = base64.b64decode(data_b64).decode('utf-8', errors='ignore')
                js = json.loads(data)
                node = {
                    "name": js.get("ps", "vmess"),
                    "type": "vmess",
                    "server": js.get("add"),
                    "port": int(js.get("port", 443)),
                    "uuid": js.get("id"),
                    "alterId": int(js.get("aid", 0)),
                    "cipher": "auto",
                    "tls": True if js.get("tls") == "tls" else False,
                    "network": js.get("net", "tcp")
                }
                if js.get("net") == "ws":
                    node["ws-opts"] = {"path": js.get("path", "/"), "headers": {"Host": js.get("host", js.get("add"))}}
                
            elif line.startswith("vless://"):
                parsed = urlparse(line)
                user_info = parsed.netloc.split("@")
                uuid = user_info[0]
                server_port = user_info[1].split(":")
                params = parse_qs(parsed.query)
                node = {
                    "name": re.sub(r'[^\w\s-]', '', urllib.parse.unquote(parsed.fragment)) if parsed.fragment else "vless",
                    "type": "vless",
                    "server": server_port[0],
                    "port": int(server_port[1]),
                    "uuid": uuid,
                    "cipher": "none",
                    "tls": True if params.get("security", [""])[0] in ["tls", "reality"] else False,
                    "network": params.get("type", ["tcp"])[0]
                }
            
            elif line.startswith("ss://"):
                parsed = urlparse(line)
                node = {
                    "name": urllib.parse.unquote(parsed.fragment) if parsed.fragment else "ss",
                    "type": "ss",
                    "server": parsed.hostname,
                    "port": parsed.port,
                    "cipher": "aes-256-gcm",
                    "password": "password"
                }

            if node:
                # 确保名称唯一性，防止 Clash 启动报错
                name = node["name"]
                counter = 1
                while name in seen_names:
                    name = f"{node['name']} ({counter})"
                    counter += 1
                node["name"] = name
                seen_names.add(name)
                proxies.append(node)
        except:
            continue

    if not proxies:
        return

    # Clash 完整配置模板
    config = {
        "port": 7890,
        "socks-port": 7891,
        "allow-lan": True,
        "mode": "rule",
        "log-level": "info",
        "proxies": proxies,
        "proxy-groups": [
            {"name": "🚀 节点选择", "type": "select", "proxies": ["♻️ 自动选择", "🎯 全球均衡"] + [p["name"] for p in proxies]},
            {"name": "♻️ 自动选择", "type": "url-test", "url": "http://www.gstatic.com/generate_204", "interval": 300, "proxies": [p["name"] for p in proxies]},
            {"name": "🎯 全球均衡", "type": "load-balance", "url": "http://www.gstatic.com/generate_204", "interval": 300, "proxies": [p["name"] for p in proxies]},
        ],
        "rules": [
            "DOMAIN-SUFFIX,google.com,🚀 节点选择",
            "DOMAIN-SUFFIX,github.com,🚀 节点选择",
            "DOMAIN-SUFFIX,telegram.org,🚀 节点选择",
            "MATCH,🚀 节点选择"
        ]
    }

    try:
        with open(CLASH_OUTPUT_FILE, "w", encoding="utf-8") as f:
            yaml.dump(config, f, allow_unicode=True, sort_keys=False)
        log(f"[*] Clash 配置文件已生成: {CLASH_OUTPUT_FILE}")
    except Exception as e:
        log(f"[-] Clash 配置生成失败: {e}")



# ============ 主处理函数 ============
def process_site(row_num, name, url, panel_type_hint):
    session = create_session()

    real_url = extract_actual_url(url, session)
    base_url = extract_base_url(real_url)
    invite_code = extract_invite_code(real_url)
    domain = urlparse(base_url).netloc
    headers = get_headers(base_url)
    panel_type = "unknown"

    log(f"\n{'='*60}")
    log(f"处理 [{row_num}] {name} - {base_url}")
    if invite_code:
        log(f"  邀请码: {invite_code}")

    # 快速连通性检测
    if not quick_connectivity_check(base_url, session):
        log(f"  站点不可达，跳过")
        return {
            "row": row_num, "name": name, "url": url, "base_url": base_url,
            "status": "跳过-不可达", "account": REG_EMAIL, "password": REG_PASSWORD,
            "subscribe_url": "", "subscribe_domain": "", "note": "连接失败或超时",
            "plans": [],
        }

    # 优先探测 V2Board 公开配置 + env.js，快速发现真实 API
    is_v2board, v2_config = probe_v2board_config(base_url, session, headers)
    if is_v2board:
        panel_type = "v2board"
    else:
        env_api = detect_api_base_from_env_js(base_url, session, headers)
        if env_api:
            base_url = env_api
            headers = get_headers(base_url)

    # 选择邮箱: 根据站点配置选择临时邮箱或 Gmail
    email_info = get_registration_email(session, v2_config)
    reg_email = email_info["email"]
    log(f"  使用邮箱: {reg_email} ({email_info['service']})")

    result = {
        "row": row_num, "name": name, "url": url, "base_url": base_url,
        "status": "", "account": reg_email, "password": REG_PASSWORD,
        "subscribe_url": "", "subscribe_domain": "", "note": "",
        "plans": [],
    }

    # 探测面板类型
    if panel_type == "unknown":
        panel_type = detect_panel_type(base_url, session, headers)
    log(f"  面板类型: {panel_type}")

    if panel_type == "cloudflare":
        result["status"] = "跳过-Cloudflare"
        result["note"] = "Cloudflare 防护"
        return result

    # Step 1: 尝试登录 (先用探测到的面板类型，失败则用备选)
    login_result = try_api_login(base_url, headers, session, reg_email, panel_type)
    if login_result["status"] == "no_api" and panel_type != "unknown":
        fallback = "v2board" if panel_type == "sspanel" else "sspanel"
        login_result2 = try_api_login(base_url, headers, session, reg_email, fallback)
        if login_result2["status"] != "no_api":
            login_result = login_result2
            panel_type = fallback
            log(f"  面板类型修正为: {panel_type}")

    # 如果所有标准端点都失败，尝试从 JS 中发现真实 API 地址
    if login_result["status"] == "no_api":
        alt_api_base = detect_api_base_from_js(base_url, session, headers)
        if alt_api_base:
            alt_headers = get_headers(alt_api_base)
            login_result = try_api_login(alt_api_base, alt_headers, session, reg_email, "v2board")
            if login_result["status"] != "no_api":
                base_url = alt_api_base
                headers = alt_headers
                log(f"  使用JS发现的API: {alt_api_base}")

    log(f"  登录结果: {login_result['status']}")

    if login_result["status"] == "cloudflare":
        result["status"] = "跳过-Cloudflare"
        result["note"] = "Cloudflare 防护"
        return result

    if login_result["status"] == "login_ok":
        result["status"] = "已注册-登录成功"
        token = login_result.get("token")
        sub = extract_subscription(base_url, headers, session, token)
        if sub:
            result["subscribe_url"] = sub["subscribe_url"]
            result["subscribe_domain"] = sub["subscribe_domain"]
            log(f"  订阅域名: {sub['subscribe_domain']}")
            fetch_and_save_nodes(sub["subscribe_url"], name)
        result["plans"] = fetch_plans(base_url, headers, session, token, panel_type)
        return result

    # Step 2: 尝试注册 (带邮箱轮询重试机制)
    max_email_retries = 3
    for email_attempt in range(max_email_retries):
        if email_attempt > 0:
            email_info = get_registration_email(session, v2_config)
            reg_email = email_info["email"]
            log(f"  🔄 邮箱被拒，尝试使用新邮箱: {reg_email} (尝试 {email_attempt+1}/{max_email_retries})")

        reg_result = try_api_register(base_url, headers, session, reg_email, invite_code, panel_type)
        if reg_result["status"] == "no_api" and panel_type != "unknown":
            fallback = "v2board" if panel_type == "sspanel" else "sspanel"
            reg_result2 = try_api_register(base_url, headers, session, reg_email, invite_code, fallback)
            if reg_result2["status"] != "no_api":
                reg_result = reg_result2
                panel_type = fallback
        log(f"  注册结果: {reg_result['status']}")

        # SSPanel 服务条款 -> 带 agree 重试
        if reg_result["status"] == "need_tos":
            log(f"  需要同意服务条款，重试...")
            tos_result = try_register_with_tos_retry(base_url, headers, session, reg_email, invite_code, reg_result.get("endpoint", ""), panel_type)
            if tos_result["status"] in ["register_ok", "need_email_code"]:
                reg_result = tos_result

        # 需要邮件验证码
        if reg_result["status"] == "need_email_code":
            log(f"  需要邮件验证码，启动验证码流程...")
            code_result = try_register_with_email_code(
                base_url, headers, session, reg_email, invite_code, reg_result.get("endpoint", ""), panel_type, email_info=email_info
            )
            reg_result = code_result  # 覆盖外层状态，处理 register_ok / email_rejected / fail

        # 判断是否因为邮箱后缀/临时邮箱被拒绝
        if reg_result["status"] == "email_rejected":
            log(f"  ❌ 站点拒绝了该邮箱，准备更换邮箱服务...")
            if email_info and email_info.get("service") not in ("gmail", None):
                temp_email_mgr.mark_failed(email_info["service"])
            continue

        # 不是邮箱本身问题，退出轮询执行后续逻辑
        break

    # ================= 处理最终注册结果 =================
    if reg_result["status"] == "cloudflare":
        result["status"] = "跳过-Cloudflare"
        result["note"] = "Cloudflare 防护"
        return result

    if reg_result["status"] == "already_registered":
        result["status"] = "已注册-密码不对"
        result["note"] = "邮箱已注册"
        return result

    if reg_result["status"] == "need_invite":
        result["status"] = "注册失败-需邀请码"
        result["note"] = "需要有效邀请码"
        return result

    if reg_result["status"] == "need_captcha":
        result["status"] = "跳过-需要人机验证"
        result["note"] = "检测到图形验证码或防刷系统"
        return result

    if reg_result["status"] == "closed":
        result["status"] = "跳过-关闭注册"
        result["note"] = "站点已关闭注册"
        return result

    if reg_result["status"] == "email_rejected":
        result["status"] = "注册失败-邮箱不支持"
        result["note"] = "所有备用邮箱均被该站点拒绝"
        return result

    if reg_result["status"] == "send_code_fail":
        result["status"] = "注册失败-发验证码失败"
        result["note"] = "请求发送验证码接口失败/被拦截"
        return result

    if reg_result["status"] == "no_code_received":
        result["status"] = "注册失败-收不到验证码"
        result["note"] = "临时邮箱/Gmail均未收到验证码"
        return result

    if reg_result["status"] == "register_fail":
        result["status"] = "注册失败"
        result["note"] = f"接口返回错误: {str(reg_result.get('data', ''))[:50]}"
        return result

    if reg_result["status"] == "register_ok":
        result["status"] = "注册成功"
        sub = extract_subscription(base_url, headers, session, reg_result.get("token"))
        if sub:
            result["subscribe_url"] = sub["subscribe_url"]
            result["subscribe_domain"] = sub["subscribe_domain"]
            fetch_and_save_nodes(sub["subscribe_url"], name)
        result["plans"] = fetch_plans(base_url, headers, session, reg_result.get("token"), panel_type)
        return result

    if reg_result["status"] == "no_api":
        result["status"] = "跳过-无API"
        result["note"] = "未找到可用API端点"
        return result

    result["status"] = f"注册失败"
    result["note"] = reg_result.get("status", "unknown")
    return result

# ============ 读取/保存 CSV 数据 ============
def load_sites_db():
    sites = []
    if not os.path.exists(CSV_PATH):
        return sites
    with open(CSV_PATH, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row_num, row in enumerate(reader, start=0):
            url = row.get("url", "").strip()
            if not url or not url.startswith("http"):
                continue
            sites.append({
                "row": row_num,
                "name": row.get("name", "").strip(),
                "url": url,
                "panel_type": row.get("panel_type", "").strip(),
                "fail_count": int(row.get("fail_count", 0) if row.get("fail_count") else 0),
                "last_status": row.get("last_status", "")
            })
    return sites

def save_sites_db(sites):
    if not sites:
        return
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["name", "url", "panel_type", "fail_count", "last_status"])
        writer.writeheader()
        for s in sites:
            writer.writerow({
                "name": s["name"],
                "url": s["url"],
                "panel_type": s["panel_type"],
                "fail_count": s["fail_count"],
                "last_status": s["last_status"]
            })


def main():
    log("=" * 60)
    log("批量机场面板自动注册脚本启动 (v2 进化版)")
    log(f"源数据库: {CSV_PATH}")
    log(f"注册邮箱: {REG_EMAIL}")
    log(f"验证码邮箱: {IMAP_EMAIL}")
    log("=" * 60)

    all_db_sites = load_sites_db()
    
    # 筛选剔除那些连续测试失败满 3 次的死链站点
    alive_sites = []
    dead_sites = []
    for s in all_db_sites:
        if s["fail_count"] >= 1:
            dead_sites.append(s)
        else:
            alive_sites.append(s)
            
    if dead_sites:
        log(f"清理死链: 本次自动淘汰了 {len(dead_sites)} 个已失效机场。")
        
    sites = alive_sites
    log(f"共加载 {len(sites)} 个可用站点准备打码测算")

    if START_FROM > 0:
        sites = sites[START_FROM:]
        log(f"从第 {START_FROM + 1} 个站点开始处理 (跳过前 {START_FROM} 个)")
    if TEST_LIMIT > 0:
        sites = sites[:TEST_LIMIT]
        log(f"测试模式: 只处理前 {TEST_LIMIT} 个站点")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "注册结果"
    out_ws.append(["序号", "名称", "网站地址", "状态", "注册邮箱", "注册密码", "订阅地址", "订阅域名", "备注", "套餐概要"])

    stats = {"total": len(sites), "success": 0, "already_registered": 0,
             "cloudflare": 0, "failed": 0, "skipped": 0}
    all_plans_data = []
    processed_count = 0

    def task_wrapper(i, site):
        nonlocal processed_count
        try:
            result = process_site(site["row"], site["name"], site["url"], site["panel_type"])
            
            with stats_lock:
                # 写入 Excel
                plan_summary = ""
                if result.get("plans"):
                    cheapest = min(result["plans"], key=lambda x: x["monthly_price"])
                    plan_summary = f"最低{cheapest['monthly_price']}元/月 {cheapest['traffic_gb']}GB"
                    all_plans_data.append({"site_name": result["name"], "plans": result["plans"]})

                out_ws.append([
                    i + 1, result["name"], result["url"], result["status"],
                    result["account"], result["password"],
                    result["subscribe_url"], result["subscribe_domain"], result["note"],
                    plan_summary,
                ])

                status = result["status"]
                if "注册成功" in status:
                    stats["success"] += 1
                elif "已注册" in status:
                    stats["already_registered"] += 1
                elif "Cloudflare" in status:
                    stats["cloudflare"] += 1
                elif "跳过" in status:
                    stats["skipped"] += 1
                else:
                    stats["failed"] += 1

                # 数据库状态维护
                healthy_keywords = ["已注册", "注册成功", "Cloudflare", "人机验证", "关闭注册"]
                if any(kw in status for kw in healthy_keywords):
                    site["fail_count"] = 0
                else:
                    site["fail_count"] += 1
                site["last_status"] = status
                
                processed_count += 1
                if processed_count % 5 == 0:
                    out_wb.save(OUTPUT_EXCEL)
                    log(f"  [进度] 已处理: {processed_count}/{len(sites)} | 成功: {stats['success']} | 失败/跳过: {stats['failed'] + stats['skipped']}")
            
            return result.get("subscribe_url", "")
        except Exception as e:
            log(f"  线程处理异常: {e}")
            return ""

    log(f"[*] 启动多线程引擎 (并发数: {MAX_WORKERS})...")
    all_subscribe_urls = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(task_wrapper, i, site) for i, site in enumerate(sites)]
        for future in as_completed(futures):
            url = future.result()
            if url:
                all_subscribe_urls.append(url)

    out_wb.save(OUTPUT_EXCEL)
    
    # 彻底清理死链并回写 CSV
    final_alive_sites = [s for s in all_db_sites if s.get("fail_count", 0) < 1]
    save_sites_db(final_alive_sites)
    
    # 生成套餐报表
    generate_plan_analysis(all_plans_data)
    
    # 生成 Clash 订阅汇总配置 (并进行全局节点去重)
    all_nodes_collected = ""
    if os.path.exists(NODES_OUTPUT_FILE):
        with open(NODES_OUTPUT_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        
        # 对 all_nodes.txt 本身进行去重写回
        unique_nodes = []
        seen_nodes = set()
        for l in lines:
            l_strip = l.strip()
            if "://" in l_strip and l_strip not in seen_nodes:
                unique_nodes.append(l_strip)
                seen_nodes.add(l_strip)
            elif l_strip.startswith("#"):
                unique_nodes.append(l_strip) # 保留注释
        
        all_nodes_collected = "\n".join(unique_nodes)
        with open(NODES_OUTPUT_FILE, "w", encoding="utf-8") as f:
            f.write(all_nodes_collected + "\n")
            
    generate_clash_config(all_nodes_collected)

    log("\n" + "=" * 60)
    log("所有任务并发处理完成!")
    log(f"  总计规模: {stats['total']}")
    log(f"  注册/获取成功: {stats['success'] + stats['already_registered']}")
    log(f"  存活数据库已精简至: {len(final_alive_sites)} 个站点")
    log(f"  最终成品: {OUTPUT_EXCEL}, {CLASH_OUTPUT_FILE}")
    log("=" * 60)


if __name__ == "__main__":
    main()
